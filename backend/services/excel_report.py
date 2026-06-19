import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from services.report_service import (
    generate_report_data,
    get_department_headcount_rows,
    get_leave_summary_rows,
    get_performance_summary_rows,
    ensure_reports_dir,
    REPORTS_DIR
)


def _style_header(cell):

    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill(
        start_color="0891B2",
        end_color="0891B2",
        fill_type="solid"
    )
    cell.alignment = Alignment(horizontal="center")


def _write_dataframe(ws, df, start_row=1):

    for col_idx, column in enumerate(df.columns, 1):
        cell = ws.cell(row=start_row, column=col_idx)
        cell.value = column
        _style_header(cell)

    for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx).value = value


def create_excel_report():

    ensure_reports_dir()

    wb = Workbook()

    data = generate_report_data()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"

    summary_rows = [
        ("HR Report", ""),
        ("Generated On", data["generated_on"]),
        ("", ""),
        ("Total Employees", data["total_employees"]),
        ("Total Departments", data["total_departments"]),
        ("Average Rating", data["average_rating"]),
        ("Average Present Days", data["average_present_days"]),
        ("Total Leave Days", data["total_leave_days"]),
        ("Average Leave Days", data["average_leave_days"]),
        ("Missing PAN Records", data["missing_pan"]),
        ("Missing NDA Records", data["missing_nda"]),
        ("Highest Performance Rating", data["performance_summary"]["highest_rating"]),
        ("Lowest Performance Rating", data["performance_summary"]["lowest_rating"]),
        ("Employees Reviewed", data["performance_summary"]["employees_reviewed"]),
    ]

    ws_summary.cell(row=1, column=1).value = "HR Analytics Report"
    ws_summary.cell(row=1, column=1).font = Font(bold=True, size=14)

    for row_idx, (label, value) in enumerate(summary_rows[1:], 3):
        ws_summary.cell(row=row_idx, column=1).value = label
        ws_summary.cell(row=row_idx, column=1).font = Font(bold=True)
        ws_summary.cell(row=row_idx, column=2).value = value

    ws_summary.column_dimensions["A"].width = 32
    ws_summary.column_dimensions["B"].width = 24

    # Department headcount
    ws_dept = wb.create_sheet("Department Headcount")
    _write_dataframe(ws_dept, get_department_headcount_rows())

    # Leave summary
    ws_leave = wb.create_sheet("Leave Summary")
    _write_dataframe(ws_leave, get_leave_summary_rows())

    # Performance
    ws_perf = wb.create_sheet("Top Performers")
    _write_dataframe(ws_perf, get_performance_summary_rows())

    path = os.path.join(REPORTS_DIR, "hr_report.xlsx")

    wb.save(path)

    return path
